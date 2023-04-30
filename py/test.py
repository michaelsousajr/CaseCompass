import openpyxl

workbook = openpyxl.load_workbook('model_sheet.xlsx')

worksheet = workbook['Working']

cell_value1 = worksheet.cell(row=1, column=1).value

print(cell_value1)


# Iterate over all rows in the worksheet
for row in worksheet.iter_rows():
    # Iterate over all cells in the row
    for cell in row:
        # Print the value of the cell
        print(cell.value)

