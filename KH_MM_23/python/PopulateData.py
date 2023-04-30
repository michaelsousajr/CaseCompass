import openpyxl
import random
import os

num = 300

# Create a new workbook
workbook = openpyxl.Workbook()

# Select the first worksheet and set its name
sheet_Data = workbook.active
sheet_Data.title = 'Working'

# Create Weights
sheet_Weights = workbook.create_sheet(title='Weights')

# Define the headers
headers = [
  'Case ID', 'Case Name', 'AI Generated Description',
  'Expected Payout for Injured', 'Injured Name', 'Offender Name(s)',
  'Injured Medical Expenses (pre-settlement)',
  'Expected Medical Expenses (post settlement)',
  'Injured Received Surgury? (Y-1/N-0)', 'Catastrophic? (Y-1/N-0)',
  'Injured Lost Wages', 'Injured Insurance Policy Value',
  'Offender(s) Insurance Policy Value', 'Police Report?',
  'Injured Guilt Liability (0 completely innocent - 4 very guilty) [synthesis of police report]',
  'Offender Guilt Liability (0 completely innocent - 4 very guilty) [synthesis of police report]',
  'Injured FL Local?', 'Offender FL Local?', 'Injured Age:',
  '(Avg)Offender(s) Age:', 'Injured Phone Number',
  'Incident Location Type (Y-Car N-Non Car)',
  'Injured Group?', 'Offender Group?', 'A coefficient', 'I coefficient',
  'E coefficient'
]

# Write the headers to the first row of the worksheet
for i in range(len(headers)):
  cell = sheet_Data.cell(row=1, column=i + 1)
  cell.value = headers[i]

## individual, car crash, mid level injury, no death, not at fault, middle class, Local
for i in range(num):
  # insert row at 2
  sheet_Data.insert_rows(2)

  # populate data
  #sheet_Data.cell(row=2, column=1).value = i      # Case ID (ignore)
  #sheet_Data.cell(row=2, column=2).value = '-'    # Case Name (ignore)
  #sheet_Data.cell(row=2, column=3).value = '-'    # Description (ignore)
  #sheet_Data.cell(row=2, column=4).value = ''      # Expected Payout (ignore)
  #sheet_Data.cell(row=2, column=5).value = ''    # Injured Name (ignore)
  #sheet_Data.cell(row=2, column=6).value = i     # Offender Name
  sheet_Data.cell(row=2, column=7).value = (round(random.uniform(15000, 25000),2))  #pre-settlement medical expenses
  sheet_Data.cell(row=2, column=8).value = (round(random.uniform(5000, 15000),2))  #post-settlement medical expenses
  sheet_Data.cell(row=2, column=9).value = 'Y'  #surgery?
  sheet_Data.cell(row=2, column=10).value = 'N'  #Catastrophic?
  sheet_Data.cell(row=2, column=11).value = (round(random.uniform(15000, 25000),2))  #lost wages
  sheet_Data.cell(row=2, column=12).value = (round(random.uniform(10000, 30000),2))  #Injured Insurance policy value
  sheet_Data.cell(row=2, column=13).value = (round(random.uniform(10000, 25000),2))  #Offender Insurance policy value
  sheet_Data.cell(row=2, column=14).value = 'Y'  #Police Report
  sheet_Data.cell(row=2, column=15).value = random.randint(0, 1) #injured guilt liability
  sheet_Data.cell(row=2, column=16).value = random.randint(2, 4) #offender guilt liability
  sheet_Data.cell(row=2, column=17).value = 'Y' #injured FL Local
  sheet_Data.cell(row=2, column=18).value = 'Y' #offended FL Local
  #sheet_Data.cell(row=2, column=19).value = '' #injured age
  #sheet_Data.cell(row=2, column=20).value = '' #offended age
  sheet_Data.cell(row=2, column=21).value = '123-456-7890' #injured phone number
  sheet_Data.cell(row=2, column=22).value = 'Y' # location type
  sheet_Data.cell(row=2, column=23).value = 'N' #injured group
  sheet_Data.cell(row=2, column=24).value = 'N' #offender group
  #sheet_Data.cell(row=2, column=25).value = '' #A coefficient
  #sheet_Data.cell(row=2, column=26).value = '' #I coefficient
  #sheet_Data.cell(row=2, column=27).value = '' #E coefficient


print('\n1')


## individual, car crash, high level injury, no death, not at fault, middle class, Local
for i in range(num):
  # insert row at 2
  sheet_Data.insert_rows(2)

  # populate data
  #sheet_Data.cell(row=2, column=1).value = i      # Case ID (ignore)
  #sheet_Data.cell(row=2, column=2).value = '-'    # Case Name (ignore)
  #sheet_Data.cell(row=2, column=3).value = '-'    # Description (ignore)
  #sheet_Data.cell(row=2, column=4).value = ''      # Expected Payout (ignore)
  #sheet_Data.cell(row=2, column=5).value = ''    # Injured Name (ignore)
  #sheet_Data.cell(row=2, column=6).value = i     # Offender Name
  sheet_Data.cell(row=2, column=7).value = (round(random.uniform(40000, 75000),2))  #pre-settlement medical expenses
  sheet_Data.cell(row=2, column=8).value = (round(random.uniform(50000, 100000),2))  #post-settlement medical expenses
  sheet_Data.cell(row=2, column=9).value = 'Y'  #surgery?
  sheet_Data.cell(row=2, column=10).value = 'N'  #Catastrophic?
  sheet_Data.cell(row=2, column=11).value = (round(random.uniform(15000, 25000),2))  #lost wages
  sheet_Data.cell(row=2, column=12).value = (round(random.uniform(10000, 50000),2))  #Injured Insurance policy value
  sheet_Data.cell(row=2, column=13).value = (round(random.uniform(10000, 50000),2))  #Offender Insurance policy value
  sheet_Data.cell(row=2, column=14).value = 'Y'  #Police Report
  sheet_Data.cell(row=2, column=15).value = random.randint(0, 1) #injured guilt liability
  sheet_Data.cell(row=2, column=16).value = random.randint(2, 4) #offender guilt liability
  sheet_Data.cell(row=2, column=17).value = 'Y' #injured FL Local
  sheet_Data.cell(row=2, column=18).value = 'Y' #offended FL Local
  #sheet_Data.cell(row=2, column=19).value = '' #injured age
  #sheet_Data.cell(row=2, column=20).value = '' #offended age
  sheet_Data.cell(row=2, column=21).value = '123-456-7890' #injured phone number
  sheet_Data.cell(row=2, column=22).value = 'Y' # location type
  sheet_Data.cell(row=2, column=23).value = 'N' #injured group
  sheet_Data.cell(row=2, column=24).value = 'N' #offender group
  #sheet_Data.cell(row=2, column=25).value = '' #A coefficient
  #sheet_Data.cell(row=2, column=26).value = '' #I coefficient
  #sheet_Data.cell(row=2, column=27).value = '' #E coefficient


print('\n2')


## individual, car crash, high level injury, death, not at fault, middle class, Local
for i in range(num):
  # insert row at 2
  sheet_Data.insert_rows(2)

  # populate data
  #sheet_Data.cell(row=2, column=1).value = i      # Case ID (ignore)
  #sheet_Data.cell(row=2, column=2).value = '-'    # Case Name (ignore)
  #sheet_Data.cell(row=2, column=3).value = '-'    # Description (ignore)
  #sheet_Data.cell(row=2, column=4).value = ''      # Expected Payout (ignore)
  #sheet_Data.cell(row=2, column=5).value = ''    # Injured Name (ignore)
  #sheet_Data.cell(row=2, column=6).value = i     # Offender Name
  sheet_Data.cell(row=2, column=7).value = (round(random.uniform(30000, 100000),2))  #pre-settlement medical expenses
  sheet_Data.cell(row=2, column=8).value = (round(random.uniform(0, 0),2))  #post-settlement medical expenses
  sheet_Data.cell(row=2, column=9).value = 'Y'  #surgery?
  sheet_Data.cell(row=2, column=10).value = 'Y'  #Catastrophic?
  sheet_Data.cell(row=2, column=11).value = (round(random.uniform(15000, 25000),2))  #lost wages
  sheet_Data.cell(row=2, column=12).value = (round(random.uniform(10000, 50000),2))  #Injured Insurance policy value
  sheet_Data.cell(row=2, column=13).value = (round(random.uniform(10000, 50000),2))  #Offender Insurance policy value
  sheet_Data.cell(row=2, column=14).value = 'Y'  #Police Report
  sheet_Data.cell(row=2, column=15).value = random.randint(0, 1) #injured guilt liability
  sheet_Data.cell(row=2, column=16).value = random.randint(2, 4) #offender guilt liability
  sheet_Data.cell(row=2, column=17).value = 'Y' #injured FL Local
  sheet_Data.cell(row=2, column=18).value = 'Y' #offended FL Local
  #sheet_Data.cell(row=2, column=19).value = '' #injured age
  #sheet_Data.cell(row=2, column=20).value = '' #offended age
  sheet_Data.cell(row=2, column=21).value = '123-456-7890' #injured phone number
  sheet_Data.cell(row=2, column=22).value = 'Y' # location type
  sheet_Data.cell(row=2, column=23).value = 'N' #injured group
  sheet_Data.cell(row=2, column=24).value = 'N' #offender group
  #sheet_Data.cell(row=2, column=25).value = '' #A coefficient
  #sheet_Data.cell(row=2, column=26).value = '' #I coefficient
  #sheet_Data.cell(row=2, column=27).value = '' #E coefficient


print('\n3')


## individual, domestic, low level injury,  no death, not at fault, middle class, Local
for i in range(num):
  # insert row at 2
  sheet_Data.insert_rows(2)

  # populate data
  #sheet_Data.cell(row=2, column=1).value = i      # Case ID (ignore)
  #sheet_Data.cell(row=2, column=2).value = '-'    # Case Name (ignore)
  #sheet_Data.cell(row=2, column=3).value = '-'    # Description (ignore)
  #sheet_Data.cell(row=2, column=4).value = ''      # Expected Payout (ignore)
  #sheet_Data.cell(row=2, column=5).value = ''    # Injured Name (ignore)
  #sheet_Data.cell(row=2, column=6).value = i     # Offender Name
  sheet_Data.cell(row=2, column=7).value = (round(random.uniform(500, 10000),2))  #pre-settlement medical expenses
  sheet_Data.cell(row=2, column=8).value = (round(random.uniform(500, 2500),2))  #post-settlement medical expenses
  sheet_Data.cell(row=2, column=9).value = 'Y'  #surgery?
  sheet_Data.cell(row=2, column=10).value = 'N'  #Catastrophic?
  sheet_Data.cell(row=2, column=11).value = (round(random.uniform(1000, 5000),2))  #lost wages
  sheet_Data.cell(row=2, column=12).value = (round(random.uniform(10000, 30000),2))  #Injured Insurance policy value
  sheet_Data.cell(row=2, column=13).value = (round(random.uniform(10000, 30000),2))  #Offender Insurance policy value
  sheet_Data.cell(row=2, column=14).value = 'Y'  #Police Report
  sheet_Data.cell(row=2, column=15).value = random.randint(0, 1) #injured guilt liability
  sheet_Data.cell(row=2, column=16).value = random.randint(2, 4) #offender guilt liability
  sheet_Data.cell(row=2, column=17).value = 'Y' #injured FL Local
  sheet_Data.cell(row=2, column=18).value = 'Y' #offended FL Local
  #sheet_Data.cell(row=2, column=19).value = '' #injured age
  #sheet_Data.cell(row=2, column=20).value = '' #offended age
  sheet_Data.cell(row=2, column=21).value = '123-456-7890' #injured phone number
  sheet_Data.cell(row=2, column=22).value = '2' # location type
  sheet_Data.cell(row=2, column=23).value = 'N' #injured group
  sheet_Data.cell(row=2, column=24).value = 'N' #offender group
  #sheet_Data.cell(row=2, column=25).value = '' #A coefficient
  #sheet_Data.cell(row=2, column=26).value = '' #I coefficient
  #sheet_Data.cell(row=2, column=27).value = '' #E coefficient


print('\n4')


## group, domestic, low level injury, no death, not at fault, middle class, Local
for i in range(num):
  # insert row at 2
  sheet_Data.insert_rows(2)

  # populate data
  #sheet_Data.cell(row=2, column=1).value = i      # Case ID (ignore)
  #sheet_Data.cell(row=2, column=2).value = '-'    # Case Name (ignore)
  #sheet_Data.cell(row=2, column=3).value = '-'    # Description (ignore)
  #sheet_Data.cell(row=2, column=4).value = ''      # Expected Payout (ignore)
  #sheet_Data.cell(row=2, column=5).value = ''    # Injured Name (ignore)
  #sheet_Data.cell(row=2, column=6).value = i     # Offender Name
  sheet_Data.cell(row=2, column=7).value = (round(random.uniform(500, 10000),2))  #pre-settlement medical expenses
  sheet_Data.cell(row=2, column=8).value = (round(random.uniform(500, 2500),2))  #post-settlement medical expenses
  sheet_Data.cell(row=2, column=9).value = 'Y'  #surgery?
  sheet_Data.cell(row=2, column=10).value = 'N'  #Catastrophic?
  sheet_Data.cell(row=2, column=11).value = (round(random.uniform(1000, 5000),2))  #lost wages
  sheet_Data.cell(row=2, column=12).value = (round(random.uniform(10000, 30000),2))  #Injured Insurance policy value
  sheet_Data.cell(row=2, column=13).value = (round(random.uniform(10000, 30000),2))  #Offender Insurance policy value
  sheet_Data.cell(row=2, column=14).value = 'Y'  #Police Report
  sheet_Data.cell(row=2, column=15).value = random.randint(0, 1) #injured guilt liability
  sheet_Data.cell(row=2, column=16).value = random.randint(2, 4) #offender guilt liability
  sheet_Data.cell(row=2, column=17).value = 'Y' #injured FL Local
  sheet_Data.cell(row=2, column=18).value = 'Y' #offended FL Local
  #sheet_Data.cell(row=2, column=19).value = '' #injured age
  #sheet_Data.cell(row=2, column=20).value = '' #offended age
  sheet_Data.cell(row=2, column=21).value = '123-456-7890' #injured phone number
  sheet_Data.cell(row=2, column=22).value = 'Y' # location type
  sheet_Data.cell(row=2, column=23).value = 'Y' #injured group
  sheet_Data.cell(row=2, column=24).value = 'Y' #offender group
  #sheet_Data.cell(row=2, column=25).value = '' #A coefficient
  #sheet_Data.cell(row=2, column=26).value = '' #I coefficient
  #sheet_Data.cell(row=2, column=27).value = '' #E coefficient


print('\n5')


## group, domestic, mid level injury, no death, kinda guilty, middle class, not car, Local
for i in range(num):
  # insert row at 2
  sheet_Data.insert_rows(2)

  # populate data
  #sheet_Data.cell(row=2, column=1).value = i      # Case ID (ignore)
  #sheet_Data.cell(row=2, column=2).value = '-'    # Case Name (ignore)
  #sheet_Data.cell(row=2, column=3).value = '-'    # Description (ignore)
  #sheet_Data.cell(row=2, column=4).value = ''      # Expected Payout (ignore)
  #sheet_Data.cell(row=2, column=5).value = ''    # Injured Name (ignore)
  #sheet_Data.cell(row=2, column=6).value = i     # Offender Name
  sheet_Data.cell(row=2, column=7).value = (round(random.uniform(2000, 15000),2))  #pre-settlement medical expenses
  sheet_Data.cell(row=2, column=8).value = (round(random.uniform(500, 5000),2))  #post-settlement medical expenses
  sheet_Data.cell(row=2, column=9).value = 'Y'  #surgery?
  sheet_Data.cell(row=2, column=10).value = 'N'  #Catastrophic?
  sheet_Data.cell(row=2, column=11).value = (round(random.uniform(1500, 10000),2))  #lost wages
  sheet_Data.cell(row=2, column=12).value = (round(random.uniform(10000, 30000),2))  #Injured Insurance policy value
  sheet_Data.cell(row=2, column=13).value = (round(random.uniform(10000, 30000),2))  #Offender Insurance policy value
  sheet_Data.cell(row=2, column=14).value = 'Y'  #Police Report
  sheet_Data.cell(row=2, column=15).value = random.randint(1, 3) #injured guilt liability
  sheet_Data.cell(row=2, column=16).value = random.randint(2, 4) #offender guilt liability
  sheet_Data.cell(row=2, column=17).value = 'Y' #injured FL Local
  sheet_Data.cell(row=2, column=18).value = 'Y' #offended FL Local
  #sheet_Data.cell(row=2, column=19).value = '' #injured age
  #sheet_Data.cell(row=2, column=20).value = '' #offended age
  sheet_Data.cell(row=2, column=21).value = '123-456-7890' #injured phone number
  sheet_Data.cell(row=2, column=22).value = 'N' # location type
  sheet_Data.cell(row=2, column=23).value = 'Y' #injured group
  sheet_Data.cell(row=2, column=24).value = 'Y' #offender group
  #sheet_Data.cell(row=2, column=25).value = '' #A coefficient
  #sheet_Data.cell(row=2, column=26).value = '' #I coefficient
  #sheet_Data.cell(row=2, column=27).value = '' #E coefficient


print("\n6")


## group, domestic, mid level injury, no death, very guilty, middle class, not car, Local
for i in range(num):
  # insert row at 2
  sheet_Data.insert_rows(2)

  # populate data
  #sheet_Data.cell(row=2, column=1).value = i      # Case ID (ignore)
  #sheet_Data.cell(row=2, column=2).value = '-'    # Case Name (ignore)
  #sheet_Data.cell(row=2, column=3).value = '-'    # Description (ignore)
  #sheet_Data.cell(row=2, column=4).value = ''      # Expected Payout (ignore)
  #sheet_Data.cell(row=2, column=5).value = ''    # Injured Name (ignore)
  #sheet_Data.cell(row=2, column=6).value = i     # Offender Name
  sheet_Data.cell(row=2, column=7).value = (round(random.uniform(2000, 15000),2))  #pre-settlement medical expenses
  sheet_Data.cell(row=2, column=8).value = (round(random.uniform(500, 5000),2))  #post-settlement medical expenses
  sheet_Data.cell(row=2, column=9).value = 'Y'  #surgery?
  sheet_Data.cell(row=2, column=10).value = 'N'  #Catastrophic?
  sheet_Data.cell(row=2, column=11).value = (round(random.uniform(1500, 10000),2))  #lost wages
  sheet_Data.cell(row=2, column=12).value = (round(random.uniform(10000, 30000),2))  #Injured Insurance policy value
  sheet_Data.cell(row=2, column=13).value = (round(random.uniform(10000, 30000),2))  #Offender Insurance policy value
  sheet_Data.cell(row=2, column=14).value = 'Y'  #Police Report
  sheet_Data.cell(row=2, column=15).value = random.randint(3, 4) #injured guilt liability
  sheet_Data.cell(row=2, column=16).value = random.randint(1, 3) #offender guilt liability
  sheet_Data.cell(row=2, column=17).value = 'Y' #injured FL Local
  sheet_Data.cell(row=2, column=18).value = 'Y' #offended FL Local
  #sheet_Data.cell(row=2, column=19).value = '' #injured age
  #sheet_Data.cell(row=2, column=20).value = '' #offended age
  sheet_Data.cell(row=2, column=21).value = '123-456-7890' #injured phone number
  sheet_Data.cell(row=2, column=22).value = 'N' # location type
  sheet_Data.cell(row=2, column=23).value = 'Y' #injured group
  sheet_Data.cell(row=2, column=24).value = 'Y' #offender group
  #sheet_Data.cell(row=2, column=25).value = '' #A coefficient
  #sheet_Data.cell(row=2, column=26).value = '' #I coefficient
  #sheet_Data.cell(row=2, column=27).value = '' #E coefficient


print("\n7")


## individual, out of state, high level injury, death, very guilty, middle class, not car, Local
for i in range(num):
  # insert row at 2
  sheet_Data.insert_rows(2)

  # populate data
  #sheet_Data.cell(row=2, column=1).value = i      # Case ID (ignore)
  #sheet_Data.cell(row=2, column=2).value = '-'    # Case Name (ignore)
  #sheet_Data.cell(row=2, column=3).value = '-'    # Description (ignore)
  #sheet_Data.cell(row=2, column=4).value = ''      # Expected Payout (ignore)
  #sheet_Data.cell(row=2, column=5).value = ''    # Injured Name (ignore)
  #sheet_Data.cell(row=2, column=6).value = i     # Offender Name
  sheet_Data.cell(row=2, column=7).value = (round(random.uniform(30000, 100000),2))  #pre-settlement medical expenses
  sheet_Data.cell(row=2, column=8).value = (round(random.uniform(1000, 10000),2))  #post-settlement medical expenses
  sheet_Data.cell(row=2, column=9).value = 'Y'  #surgery?
  sheet_Data.cell(row=2, column=10).value = 'Y'  #Catastrophic?
  sheet_Data.cell(row=2, column=11).value = (round(random.uniform(10000, 100000),2))  #lost wages
  sheet_Data.cell(row=2, column=12).value = (round(random.uniform(10000, 100000),2))  #Injured Insurance policy value
  sheet_Data.cell(row=2, column=13).value = (round(random.uniform(10000, 100000),2))  #Offender Insurance policy value
  sheet_Data.cell(row=2, column=14).value = 'Y'  #Police Report
  sheet_Data.cell(row=2, column=15).value = random.randint(3, 4) #injured guilt liability
  sheet_Data.cell(row=2, column=16).value = random.randint(0, 2) #offender guilt liability
  sheet_Data.cell(row=2, column=17).value = 'N' #injured FL Local
  sheet_Data.cell(row=2, column=18).value = 'N' #offended FL Local
  #sheet_Data.cell(row=2, column=19).value = '' #injured age
  #sheet_Data.cell(row=2, column=20).value = '' #offended age
  sheet_Data.cell(row=2, column=21).value = '123-456-7890' #injured phone number
  sheet_Data.cell(row=2, column=22).value = 'N' # location type
  sheet_Data.cell(row=2, column=23).value = 'N' #injured group
  sheet_Data.cell(row=2, column=24).value = 'N' #offender group
  #sheet_Data.cell(row=2, column=25).value = '' #A coefficient
  #sheet_Data.cell(row=2, column=26).value = '' #I coefficient
  #sheet_Data.cell(row=2, column=27).value = '' #E coefficient


print("\n8")


# Save the workbook to a specific folder on a Mac
folder_path = '/Users/sebastianrowe/Desktop/programming/Visual Studio Code/projects/KH_MM_23/xlsx/'
file_path = os.path.join(folder_path, 'trainingData.xlsx')
workbook.save(file_path)


print("\n\nsaved\n\n")
